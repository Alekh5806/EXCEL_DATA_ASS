from pathlib import Path
from datetime import datetime

from django.utils import timezone
from openpyxl import load_workbook

from .models import ProcessData


FIELD_MAP = {
    "PYGAS FLOW RATE": "pygas_flow_rate",
    "BIOMASS FLOW RATE": "biomass_flow_rate",
    "BIOMASS TEMPERATURE": "biomass_temperature",
    "REACTOR GAS FLOW RATE": "reactor_gas_flow_rate",
    "REACTOR GAS TEMPERATURE": "reactor_gas_temperature",
    "HEAT CARRIER FLOW RATE": "heat_carrier_flow_rate",
    "HEAT CARRIER TEMPERATURE": "heat_carrier_temperature",
    "PRODUCT GAS TEMPERATURE": "product_gas_temperature",
    "HEAT CARRIER RETURN TEMPERATURE": "heat_carrier_return_temperature",
    "COMMENTS 1\n(STAGE)": "stage",
    "COMMENTS 2\n(PROCESS NOTES)": "notes",
}


def import_excel_workbook(file_path, sheet_name=None):
    workbook_path = Path(file_path)
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    sheets = [workbook[sheet_name]] if sheet_name else workbook.worksheets

    total_created = 0
    imported_sheets = []

    for worksheet in sheets:
        created = import_worksheet(worksheet, workbook_path.name)
        if created:
            imported_sheets.append(worksheet.title)
            total_created += created

    workbook.close()

    return {
        "source_file": workbook_path.name,
        "rows_created": total_created,
        "sheets_imported": imported_sheets,
    }


def import_worksheet(worksheet, source_file):
    header_row = find_header_row(worksheet)
    if header_row is None:
        return 0

    headers = [
        str(value).strip() if value is not None else ""
        for value in next(worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    ]
    column_map = build_column_map(headers)
    data_start_row = header_row + 3
    objects = []

    for row in worksheet.iter_rows(min_row=data_start_row, values_only=True):
        timestamp = parse_timestamp(row[0] if row else None)
        if timestamp is None:
            continue

        values = {
            "timestamp": timestamp,
            "date": timestamp.date(),
            "time": timestamp.time(),
            "source_file": source_file,
        }

        for index, field_name in column_map.items():
            cell_value = row[index] if index < len(row) else None
            if field_name in {"stage", "notes"}:
                values[field_name] = parse_text(cell_value)
            else:
                values[field_name] = parse_float(cell_value)

        objects.append(ProcessData(**values))

    ProcessData.objects.bulk_create(objects, batch_size=1000)
    return len(objects)


def find_header_row(worksheet):
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        first_cell = row[0] if row else None
        if first_cell == "Description ->":
            return row_number
    return None


def build_column_map(headers):
    column_map = {}
    for index, header in enumerate(headers):
        field_name = FIELD_MAP.get(header)
        if field_name:
            column_map[index] = field_name
    return column_map


def parse_timestamp(value):
    if value is None:
        return None

    if isinstance(value, datetime):
        timestamp = value
    else:
        return None

    if timezone.is_naive(timestamp):
        timestamp = timezone.make_aware(timestamp, timezone.get_current_timezone())

    return timestamp


def parse_float(value):
    if value in ("", None):
        return None

    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_text(value):
    if value is None:
        return ""
    return str(value).strip()
